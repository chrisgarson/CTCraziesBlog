#!/usr/bin/env python3.11
"""
gen_pages_template.py — Batch processing template for CT Crazies website.

HOW TO USE FOR EACH NEW BATCH:
  1. Copy this file to /home/ubuntu/gen_pages.py
  2. Fill in the BATCH CONFIGURATION section below with the new batch details.
     - Set EXCEL_PATH to the uploaded Excel file path.
     - Set IMAGES_DIR to the folder containing the batch images.
     - Set CDN_MAP to a dict of {filename: full_cdn_url} for every image.
       (CDN URLs are printed by manus-upload-file after upload.)
     - Set NEW_PAGES to batch_size / 20 (e.g. 1 for 20 articles, 2 for 40).
  3. Run: python3.11 /home/ubuntu/gen_pages.py

This template is committed to the repository so it survives sandbox resets.
DO NOT edit the structure below the BATCH CONFIGURATION section unless
you are intentionally changing the batch processing workflow.

════════════════════════════════════════════════════════════════════════════════
  *** CRITICAL PAGE ORDERING RULE — DO NOT CHANGE — ENFORCED SINCE BATCH 1 ***

  New batch articles ALWAYS go on the Home page and Page 2 (and Page 3, 4, etc.
  if the batch is larger than 20 articles). All existing pages shift BACK by the
  number of new pages added (N = batch_size / 20).

  Example — 40-article batch (N=2):
    old Home   → Page 3
    old Page 2 → Page 4
    old Page 3 → Page 5  ... and so on
    new Home   ← articles (newest 20)
    new Page 2 ← articles (next 20)

  NEVER append new articles to the end of the page list.
  NEWEST articles are ALWAYS on the Home page.
  The shift_pages() function below implements this automatically for any N.
════════════════════════════════════════════════════════════════════════════════

────────────────────────────────────────────────────────────────────────────────
BATCH PROCESSING WORKFLOW (fully automated — no manual steps required):
  1. Upload images to CDN via manus-upload-file (WITHOUT --webdev flag)
     IMPORTANT: Do NOT use --webdev flag — it returns /manus-storage/ paths
     that only work on the deployed site. Use plain manus-upload-file to get
     full https://files.manuscdn.com/... CDN URLs that work everywhere.
  2. Fill in CDN_MAP below with the returned URLs (one entry per image file).
  3. Run: python3.11 /home/ubuntu/gen_pages.py
     The script will automatically:
       a. Read the Excel file and extract article data.
       b. Pre-flight check: verify all image filenames in Excel have a CDN URL.
       c. Shift existing pages down by N.
       d. Generate new Home.tsx (and additional pages if batch > 20).
       e. Re-count all page files on disk to get the correct total.
       f. Patch totalPages in ALL existing shifted pages.
       g. Rewrite App.tsx with updated imports and routes.
       h. Run rebuild_search.py — validates all pages and writes Search.tsx.
          - If validation FAILS: the script exits with a non-zero code and
            prints exactly which pages have the wrong article count.
            DO NOT publish until the error is resolved.
          - If validation PASSES: Search.tsx is updated with the full index.
  4. Save checkpoint, then publish.
────────────────────────────────────────────────────────────────────────────────
"""

import os
import re
import sys
import subprocess
import openpyxl
from datetime import date

# ════════════════════════════════════════════════════════════════════════════════
# BATCH CONFIGURATION — fill these in for each new batch
# ════════════════════════════════════════════════════════════════════════════════

EXCEL_PATH = '/home/ubuntu/upload/<BATCH_EXCEL_FILE>.xlsx'  # path to Excel file
IMAGES_DIR = '/home/ubuntu/upload/<BATCH_IMAGES_DIR>/'      # folder with uploaded images

# CDN_MAP: map each image filename to its full CDN URL.
# These are the URLs printed by manus-upload-file after upload.
# Example:
#   CDN_MAP = {
#       '2026-05-01_120000.jpg': 'https://files.manuscdn.com/user_upload_by_module/session_file/XXXXX/YYYYY.jpg',
#       ...
#   }
CDN_MAP = {
    # FILL IN: 'filename.jpg': 'https://files.manuscdn.com/...',
}

NEW_PAGES = 1   # number of new pages this batch adds (batch_size / 20)

PAGES_DIR      = '/home/ubuntu/x-post-platform/client/src/pages'
APP_TSX        = '/home/ubuntu/x-post-platform/client/src/App.tsx'
REBUILD_SCRIPT = '/home/ubuntu/rebuild_search.py'

# ════════════════════════════════════════════════════════════════════════════════
# STEP 0: Pre-flight checks — catch configuration errors before touching any files
# ════════════════════════════════════════════════════════════════════════════════

def preflight_check(articles):
    """
    Verify all image filenames referenced in the Excel have a CDN URL in CDN_MAP.
    Abort immediately if any are missing — before any files are modified.
    """
    errors = []

    if not CDN_MAP:
        errors.append('CDN_MAP is empty. Fill in the image filename → CDN URL mapping.')

    for art in articles:
        fname = art['imageName']
        if fname not in CDN_MAP:
            errors.append(f'  Missing CDN URL for image: {fname}')
        else:
            url = CDN_MAP[fname]
            if not url.startswith('https://files.manuscdn.com/'):
                errors.append(f'  CDN URL for {fname} does not start with https://files.manuscdn.com/: {url}')
            if url.endswith('/'):
                errors.append(f'  CDN URL for {fname} has trailing slash (remove it): {url}')

    if len(articles) % 20 != 0:
        errors.append(
            f'  Article count ({len(articles)}) is not a multiple of 20. '
            f'Each page must have exactly 20 articles.'
        )

    if errors:
        print('\n*** PRE-FLIGHT CHECK FAILED — no files have been modified ***\n')
        for e in errors:
            print(e)
        print('\nFix the issues above, then re-run the script.')
        sys.exit(1)

    print(f'  Pre-flight check passed: {len(articles)} articles, {len(CDN_MAP)} CDN URLs.')

# ════════════════════════════════════════════════════════════════════════════════
# STEP 1: Read Excel data
# ════════════════════════════════════════════════════════════════════════════════

def read_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Locate columns by header name (case-insensitive, flexible)
    def col_idx(names):
        for name in names:
            for i, h in enumerate(headers):
                if h.lower() == name.lower():
                    return i
        raise ValueError(f'Could not find column with any of these names: {names}')

    idx_num      = col_idx(['NUM', 'Num', 'num'])
    idx_headline = col_idx(['X-Post Headline', 'Headline'])
    idx_tinyurl  = col_idx(['Tiny URL', 'Tiny url', 'TinyURL', 'tinyurl'])
    idx_image    = col_idx(['ImageName', 'Image Name', 'imagename'])
    idx_xpost    = col_idx(['X-Post URL', 'X-Post Url', 'XPostURL', 'xposturl'])

    articles = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        num = row[idx_num]
        if not num:
            continue
        # Validate required fields — warn but don't crash
        headline = str(row[idx_headline]).strip() if row[idx_headline] else ''
        tinyurl  = str(row[idx_tinyurl]).strip()  if row[idx_tinyurl]  else ''
        image    = str(row[idx_image]).strip()     if row[idx_image]    else ''
        xpost    = str(row[idx_xpost]).strip()     if row[idx_xpost]    else ''
        if not headline or not tinyurl or not image or not xpost:
            print(f'  WARNING: Row NUM={num} has missing field(s) — skipping.')
            continue
        articles.append({
            'num':       int(num),
            'headline':  headline,
            'tinyUrl':   tinyurl,
            'imageName': image,
            'xPostUrl':  xpost,
        })
    # Sort descending by NUM (highest first = newest first)
    articles.sort(key=lambda a: a['num'], reverse=True)
    return articles

# ════════════════════════════════════════════════════════════════════════════════
# STEP 2: Shift existing pages down
# ════════════════════════════════════════════════════════════════════════════════

def shift_pages(new_pages_count):
    """Rename existing pages in descending order to avoid collisions."""
    existing = []
    for fname in os.listdir(PAGES_DIR):
        m = re.match(r'^Page(\d+)\.tsx$', fname)
        if m:
            existing.append(int(m.group(1)))
    existing.sort(reverse=True)  # rename highest first to avoid collisions
    for n in existing:
        src = os.path.join(PAGES_DIR, f'Page{n}.tsx')
        dst = os.path.join(PAGES_DIR, f'Page{n + new_pages_count}.tsx')
        os.rename(src, dst)
        print(f'  Renamed Page{n}.tsx → Page{n + new_pages_count}.tsx')
    # Home.tsx becomes Page{new_pages_count + 1}.tsx
    home_src = os.path.join(PAGES_DIR, 'Home.tsx')
    home_dst = os.path.join(PAGES_DIR, f'Page{new_pages_count + 1}.tsx')
    os.rename(home_src, home_dst)
    print(f'  Renamed Home.tsx → Page{new_pages_count + 1}.tsx')

# ════════════════════════════════════════════════════════════════════════════════
# STEP 3: Generate page file
# ════════════════════════════════════════════════════════════════════════════════

def escape_headline(h):
    """Escape characters that are unsafe in JSX string attributes."""
    return (h.replace('&', '&amp;')
             .replace('"', '&quot;')
             .replace("'", '&#39;')
             .replace('<', '&lt;')
             .replace('>', '&gt;'))

def make_page(articles, page_label, total_pages, is_home=False):
    """Generate a .tsx page file for the given list of articles."""
    fname = 'Home.tsx' if is_home else f'{page_label}.tsx'
    fpath = os.path.join(PAGES_DIR, fname)

    blocks = []
    for art in articles:
        img_url  = CDN_MAP[art['imageName']]   # full CDN URL from CDN_MAP
        headline = escape_headline(art['headline'])
        # IMPORTANT: attribute order must match BLOCK_RE in rebuild_search.py
        # Order: headline, tinyUrl, xPostUrl, imageSrc
        blocks.append(
            f'      <ArticleBlock\n'
            f'        headline="{headline}"\n'
            f'        tinyUrl="{art["tinyUrl"]}"\n'
            f'        xPostUrl="{art["xPostUrl"]}"\n'
            f'        imageSrc="{img_url}"\n'
            f'      />'
        )

    blocks_str = '\n'.join(blocks)
    page_num   = 1 if is_home else int(re.search(r'\d+', page_label).group())
    today      = date.today().strftime('%B %d, %Y')
    total_articles = total_pages * 20

    stats_line = ''
    if is_home:
        stats_line = f'''
      {{/* SITE STATS — auto-updated by gen_pages.py. DO NOT edit manually. */}}
      <p className="text-sm text-gray-500 mt-2" style={{{{fontFamily: 'Roboto Slab, serif'}}}}>
        Last updated: <strong style={{{{ color: '#555' }}}}>{today}</strong> &nbsp;|&nbsp; Total articles: <strong style={{{{ color: '#555' }}}}>{total_articles}</strong>
      </p>'''

    content = f'''import ArticleBlock from '../components/ArticleBlock';
import PageHeader from '../components/PageHeader';
import Pagination from '../components/Pagination';

export default function {"Home" if is_home else page_label}() {{
  return (
    <div className="max-w-4xl mx-auto px-4 py-8">
      <PageHeader />
      <div className="space-y-12">
{blocks_str}
      </div>
      <Pagination currentPage={{{page_num}}} totalPages={{{total_pages}}} />{stats_line}
    </div>
  );
}}
'''
    with open(fpath, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f'  Written {fname} ({len(articles)} articles)')

# ════════════════════════════════════════════════════════════════════════════════
# STEP 4: Patch totalPages in ALL existing shifted page files
# ════════════════════════════════════════════════════════════════════════════════

def patch_total_pages(total_pages):
    """
    Update the totalPages={N} value in every Page*.tsx and Home.tsx file.
    This ensures pagination is consistent across all pages after a batch.
    """
    patched = 0
    pattern     = re.compile(r'totalPages=\{\d+\}')
    replacement = f'totalPages={{{total_pages}}}'

    for fname in os.listdir(PAGES_DIR):
        if not (fname == 'Home.tsx' or re.match(r'^Page\d+\.tsx$', fname)):
            continue
        fpath = os.path.join(PAGES_DIR, fname)
        with open(fpath, 'r', encoding='utf-8') as f:
            content = f.read()
        new_content, count = pattern.subn(replacement, content)
        if count > 0:
            with open(fpath, 'w', encoding='utf-8') as f:
                f.write(new_content)
            patched += 1

    print(f'  Patched totalPages={{{total_pages}}} in {patched} page file(s)')

# ════════════════════════════════════════════════════════════════════════════════
# STEP 5: Rewrite App.tsx with correct imports and routes for all pages
# ════════════════════════════════════════════════════════════════════════════════

def update_app_tsx(total_pages):
    """
    Fully rewrite App.tsx so imports and routes exactly match the current
    set of pages (Home + Page2 through Page{total_pages}).
    This eliminates all manual route registration and double-brace bugs.
    """
    import_lines = ['import { Switch, Route } from "wouter";']
    import_lines.append('import Home from "./pages/Home";')
    for n in range(2, total_pages + 1):
        import_lines.append(f'import Page{n} from "./pages/Page{n}";')
    import_lines.append('import Search from "./pages/Search";')

    route_lines = ['        <Route path="/" component={Home} />']
    for n in range(2, total_pages + 1):
        route_lines.append(f'        <Route path="/page{n}" component={{Page{n}}} />')
    route_lines.append('        <Route path="/search" component={Search} />')

    imports_str = '\n'.join(import_lines)
    routes_str  = '\n'.join(route_lines)

    content = f'''{imports_str}

function App() {{
  return (
    <>
      <Switch>
{routes_str}
      </Switch>
    </>
  );
}}

export default App;
'''
    with open(APP_TSX, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f'  Rewrote App.tsx: Home + Page2–Page{total_pages} + Search ({total_pages} page routes)')

# ════════════════════════════════════════════════════════════════════════════════
# STEP 6: Run rebuild_search.py (validation + search index)
# ════════════════════════════════════════════════════════════════════════════════

def run_validation(total_pages):
    expected_total = total_pages * 20
    print('\n' + '=' * 60)
    print('  Running rebuild_search.py validation...')
    print('=' * 60)
    result = subprocess.run(
        [sys.executable, REBUILD_SCRIPT, f'--expected-total={expected_total}'],
        capture_output=False  # print output directly to terminal
    )
    if result.returncode != 0:
        print('\n*** VALIDATION FAILED — DO NOT PUBLISH UNTIL RESOLVED ***')
        print('  Search.tsx was NOT updated.')
        print('  Review the page errors above, fix the issue, and re-run this script.')
        sys.exit(1)
    else:
        print('\n  Validation passed. Search.tsx updated successfully.')
    return 0

# ════════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print(f'Reading Excel: {EXCEL_PATH}')
    articles = read_excel(EXCEL_PATH)
    print(f'  {len(articles)} articles loaded (NUMs {articles[-1]["num"]}–{articles[0]["num"]})')

    # Step 0: Pre-flight — verify CDN_MAP and article count before touching files
    print('\nRunning pre-flight checks...')
    preflight_check(articles)

    # Step 2: Shift existing pages
    print(f'\nShifting existing pages down by {NEW_PAGES}...')
    shift_pages(NEW_PAGES)

    # Step 3: Generate new pages
    pages_data = [articles[i:i+20] for i in range(0, len(articles), 20)]

    print(f'\nGenerating {NEW_PAGES} new page(s)...')
    # Use a temporary total_pages estimate for generation (will be confirmed below)
    temp_total = len([f for f in os.listdir(PAGES_DIR) if re.match(r'^Page\d+\.tsx$', f)]) + NEW_PAGES
    for i, page_articles in enumerate(pages_data):
        is_home    = (i == 0)
        page_label = 'Home' if is_home else f'Page{i + 1}'
        make_page(page_articles, page_label, temp_total, is_home=is_home)

    # Step 3b: Re-count ALL page files on disk now that Home.tsx has been written.
    # This is the authoritative total — it cannot be wrong because it reads reality.
    total_pages = len([
        f for f in os.listdir(PAGES_DIR)
        if f == 'Home.tsx' or re.match(r'^Page\d+\.tsx$', f)
    ])
    print(f'  Confirmed total pages on disk: {total_pages}')

    # Step 4: Patch totalPages across all pages (using confirmed total)
    print(f'\nPatching totalPages across all pages...')
    patch_total_pages(total_pages)

    # Step 5: Rewrite App.tsx
    print(f'\nUpdating App.tsx routes...')
    update_app_tsx(total_pages)

    print(f'\nBatch complete. Total pages: {total_pages}, Total articles: {total_pages * 20}')

    # Step 6: Validate and rebuild search index
    run_validation(total_pages)
