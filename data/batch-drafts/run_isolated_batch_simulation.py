"""Run a non-publishing 20-article CTCrazies batch simulation in a temporary copy.

This script deliberately does not call GitHub, Cloudflare Pages, or R2. It uses
a mocked R2 transport to exercise the real signing/upload receipt path without
making a network request. The canonical project files are hashed before and
after the run; any unexpected source mutation fails the simulation.
"""

from __future__ import annotations

import hashlib
import json
import os
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
REPORT_DIR = Path("/home/ubuntu/ctcrazies-simulation-reports")
sys.path.insert(0, str(ROOT))

from ctc_info_xlsx import write_tagging_draft  # noqa: E402
from safe_batch import (  # noqa: E402
    R2_IMAGE_ORIGIN,
    apply_tag_plan,
    combine_batch,
    load_ledger,
    upload_batch_images_to_r2,
    validate_batch,
    write_ledger,
    write_project_from_ledger,
)
from verify_safe_site import verify  # noqa: E402


PROTECTED_PATHS = (
    Path("data/article-ledger.json"),
    Path("client/src/App.tsx"),
    Path("client/src/pages/Search.tsx"),
    Path("client/src/data/tag-index.json"),
    Path("client/public/tag-index.json"),
    Path("client/public/search-index.json"),
)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def protected_snapshot() -> dict[str, str]:
    return {str(relative): sha256(ROOT / relative) for relative in PROTECTED_PATHS}


def write_simulation_workbook(path: Path, maximum_num: int, images_dir: Path) -> list[int]:
    numbers = list(range(maximum_num + 20, maximum_num, -1))
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet["A2"] = "SIMULATION ONLY — Do not publish or upload these records."
    worksheet["B4"] = "Num"
    worksheet["D4"] = "X-Post Headline"
    worksheet["E4"] = "Source URL"
    worksheet["F4"] = "ImageName"
    worksheet["G4"] = "X-Post Url"
    # Minimal valid JPEG. The validator only requires a present local file;
    # this content also gives the mocked R2 path deterministic bytes to sign.
    jpeg = bytes.fromhex("ffd8ffe000104a46494600010100000100010000ffdb00430000ffd9")
    for offset, number in enumerate(numbers, start=5):
        image_name = f"simulation-{number}.jpg"
        worksheet.cell(offset, 2, number)
        worksheet.cell(offset, 4, f"Simulation Workflow Validation Article NUM {number}")
        worksheet.cell(offset, 5, f"https://example.invalid/simulation/{number}")
        worksheet.cell(offset, 6, image_name)
        worksheet.cell(offset, 7, f"https://x.com/C3Heditor/status/9000000000000{number}")
        (images_dir / image_name).write_bytes(jpeg)
    workbook.save(path)
    workbook.close()
    return numbers


class MockR2Response:
    status = 200

    def __enter__(self) -> "MockR2Response":
        return self

    def __exit__(self, *_: object) -> None:
        return None


def run_mocked_r2_upload(batch: list[dict[str, Any]], receipt_path: Path) -> dict[str, Any]:
    """Exercise the actual R2 signing/receipt code without issuing a network request."""
    import safe_batch

    original_urlopen = safe_batch.urllib.request.urlopen
    previous_access = os.environ.get("R2_ACCESS_KEY_ID")
    previous_secret = os.environ.get("R2_SECRET_ACCESS_KEY")
    calls: list[str] = []

    def fake_urlopen(request: Any, timeout: int = 0) -> MockR2Response:
        assert timeout == 60
        assert request.full_url.startswith("https://ctcrazies-article-images.")
        assert request.get_method() == "PUT"
        assert request.data
        calls.append(request.full_url)
        return MockR2Response()

    try:
        os.environ["R2_ACCESS_KEY_ID"] = "simulation-access-key"
        os.environ["R2_SECRET_ACCESS_KEY"] = "simulation-secret-key"
        safe_batch.urllib.request.urlopen = fake_urlopen
        receipt = upload_batch_images_to_r2(batch, receipt_path)
    finally:
        safe_batch.urllib.request.urlopen = original_urlopen
        if previous_access is None:
            os.environ.pop("R2_ACCESS_KEY_ID", None)
        else:
            os.environ["R2_ACCESS_KEY_ID"] = previous_access
        if previous_secret is None:
            os.environ.pop("R2_SECRET_ACCESS_KEY", None)
        else:
            os.environ["R2_SECRET_ACCESS_KEY"] = previous_secret

    assert len(calls) == 20
    assert len(receipt["uploaded"]) == 20
    return receipt


def run_simulation() -> dict[str, Any]:
    before = protected_snapshot()
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(prefix="ctcrazies_batch_simulation_") as directory:
        workspace = Path(directory)
        simulated_project = workspace / "project"
        shutil.copytree(ROOT / "client", simulated_project / "client")
        (simulated_project / "data").mkdir(parents=True)
        shutil.copy2(ROOT / "data" / "article-ledger.json", simulated_project / "data" / "article-ledger.json")

        images_dir = workspace / "incoming-images"
        images_dir.mkdir()
        workbook_path = workspace / "CTC Info Simulation.xlsx"
        ledger_path = simulated_project / "data" / "article-ledger.json"
        ledger = load_ledger(ledger_path)
        pre_batch_max_num = max(article["num"] for article in ledger["articles"])
        numbers = write_simulation_workbook(workbook_path, pre_batch_max_num, images_dir)

        draft_path = workspace / "simulation-draft.json"
        metadata = write_tagging_draft(workbook_path, images_dir, draft_path)
        draft = json.loads(draft_path.read_text(encoding="utf-8"))
        tag_plan = {
            "tags": {str(number): ["Texas", "China"] for number in numbers},
            "allowFiveTagNums": [],
            "allowSixTagNums": [],
        }
        reviewed = apply_tag_plan(draft, tag_plan, ledger)
        validated = validate_batch(ledger, reviewed, workbook_path)
        r2_receipt = run_mocked_r2_upload(validated, workspace / "simulated-r2-upload-receipt.json")
        combined = combine_batch(ledger, validated, "Simulation Only — Not Published")
        write_project_from_ledger(combined, "Simulation Only — Not Published", simulated_project)
        write_ledger(combined, ledger_path)
        errors = verify(simulated_project, ledger_path)
        if errors:
            raise RuntimeError("Simulation publication gate failed: " + " | ".join(errors))

        simulated = load_ledger(ledger_path)
        expected_page_one = numbers
        actual_page_one = [article["num"] for article in simulated["articles"][:20]]
        if actual_page_one != expected_page_one:
            raise RuntimeError(f"Simulation Page 1 NUM order mismatch: {actual_page_one}")
        if simulated["articles"][20]["num"] != pre_batch_max_num:
            raise RuntimeError(
                f"Simulation Page 2 did not begin with the previous newest article, NUM {pre_batch_max_num}"
            )

        result = {
            "status": "passed",
            "simulationOnly": True,
            "incomingArticles": len(validated),
            "preBatchMaxNum": pre_batch_max_num,
            "totalArticlesAfterSimulation": len(simulated["articles"]),
            "pagesAfterSimulation": len(simulated["articles"]) // 20,
            "simulatedNums": numbers,
            "pageOneNums": actual_page_one,
            "pageTwoFirstNum": simulated["articles"][20]["num"],
            "mockedR2Uploads": len(r2_receipt["uploaded"]),
            "mockedR2Urls": [entry["imageUrl"] for entry in r2_receipt["uploaded"]],
            "canonicalR2Prefix": R2_IMAGE_ORIGIN,
            "publicationGateErrors": errors,
            "originalProjectHashesUnchanged": before == protected_snapshot(),
            "networkServicesContacted": [],
        }

    if not result["originalProjectHashesUnchanged"]:
        raise RuntimeError("Simulation altered a protected canonical project file")
    report_path = REPORT_DIR / "latest-ctcrazies-batch-simulation.json"
    report_path.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    return {**result, "reportPath": str(report_path)}


if __name__ == "__main__":
    print(json.dumps(run_simulation(), indent=2))
