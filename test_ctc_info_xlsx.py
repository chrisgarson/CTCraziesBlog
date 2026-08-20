"""Unit tests for the new CTC Info workbook reader."""

import tempfile
import unittest
from pathlib import Path

import openpyxl

from ctc_info_xlsx import read_ctc_info_workbook
from validate_batch import xlsx_cross_check


class CTCInfoWorkbookTests(unittest.TestCase):
    def test_reads_row_four_headers_and_preserves_new_source_urls(self):
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "sample.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet["A2"] = "Mandatory Read: follow the declared CTC Info processing rules."
            sheet.append([])
            sheet.append(["DateCaptrd", "Num", "WebPage#", "X-Post Headline", "Source URL", "ImageName", "X-Post Url"])
            sheet.append(["08-11-26", 1420, 1, "Exact Headline ", "https://source.example/a", "a.jpg", "https://x.com/example/a"])
            sheet.append(["08-11-26", 1419, 1, "Second Headline", "https://source.example/b", "b.jpg", "https://x.com/example/b"])
            workbook.save(xlsx)

            articles, metadata = read_ctc_info_workbook(xlsx, tmp)

            self.assertEqual(metadata["headerRow"], 4)
            self.assertEqual(metadata["dataStartRow"], 5)
            self.assertEqual(metadata["articleCount"], 2)
            self.assertEqual([article["num"] for article in articles], [1420, 1419])
            self.assertEqual(articles[0]["headline"], "Exact Headline ")
            self.assertEqual(articles[0]["sourceUrl"], "https://source.example/a")
            self.assertEqual(articles[0]["tinyUrl"], "https://source.example/a")
            self.assertEqual(articles[0]["imageName"], "a.jpg")
            self.assertEqual(articles[0]["xPostUrl"], "https://x.com/example/a")

    def test_cross_check_uses_num_and_source_url_from_new_layout(self):
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "sample.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet["A2"] = "Mandatory Read: use the CTC Info layout."
            sheet.append([])
            sheet.append(["DateCaptrd", "Num", "WebPage#", "X-Post Headline", "Source URL", "ImageName", "X-Post Url"])
            sheet.append(["08-11-26", 1420, 1, "Exact Headline", "https://source.example/a", "a.jpg", "https://x.com/example/a"])
            sheet.append(["08-11-26", 1419, 1, "Second Headline", "https://source.example/b", "b.jpg", "https://x.com/example/b"])
            workbook.save(xlsx)

            # Reversed JSON order proves the checker uses NUM, not row position.
            draft = [
                {"num": 1419, "headline": "Second Headline", "sourceUrl": "https://source.example/b"},
                {"num": 1420, "headline": "Exact Headline", "tinyUrl": "https://source.example/a"},
            ]
            self.assertEqual(xlsx_cross_check(draft, str(xlsx)), [])

    def test_resolves_a_unique_image_name_inside_a_nested_subdirectory(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            xlsx = root / "sample.xlsx"
            nested_images = root / "CTC Images"
            nested_images.mkdir()
            (nested_images / "a.jpg").write_bytes(b"image")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet["A2"] = "Mandatory Read: use the CTC Info layout."
            sheet.append([])
            sheet.append(["DateCaptrd", "Num", "WebPage#", "X-Post Headline", "Source URL", "ImageName", "X-Post Url"])
            sheet.append(["08-20-26", 1500, 1, "Exact Headline", "https://source.example/a", "a.jpg", "https://x.com/example/a"])
            workbook.save(xlsx)

            articles, _ = read_ctc_info_workbook(xlsx, root)

            self.assertEqual(articles[0]["imagePath"], str(nested_images / "a.jpg"))


if __name__ == "__main__":
    unittest.main()
