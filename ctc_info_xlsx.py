"""Read the current CTC Info XLSX layout without altering user-supplied headlines.

The current layout is deliberately positional:
  row 1: blank
  row 2: mandatory instruction note
  row 3: blank
  row 4: headers
  row 5 onward: article rows

Only D-G create article content.  Column B NUM is retained for ordering and
validation; columns A and C remain reference information.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


HEADER_ROW = 4
DATA_START_ROW = 5
REQUIRED_HEADERS = {
    "num": "Num",
    "headline": "X-Post Headline",
    "sourceUrl": "Source URL",
    "imageName": "ImageName",
    "xPostUrl": "X-Post Url",
}


def _text(value: Any) -> str:
    """Return cell text without trimming or changing the supplied wording."""
    if value is None:
        return ""
    return value if isinstance(value, str) else str(value)


def _is_blank(value: Any) -> bool:
    return not _text(value).strip()


def _header_key(value: Any) -> str:
    return "".join(ch.lower() for ch in _text(value) if ch.isalnum())


def _column_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    """Require the declared Row 4 layout rather than guessing a header row."""
    actual = {_header_key(ws.cell(HEADER_ROW, col).value): col for col in range(1, ws.max_column + 1)}
    expected = {
        "num": "num",
        "headline": "xpostheadline",
        "sourceUrl": "sourceurl",
        "imageName": "imagename",
        "xPostUrl": "xposturl",
    }
    missing = [label for key, label in expected.items() if label not in actual]
    if missing:
        raise ValueError(
            "CTC Info XLSX must have its headers on Row 4. Missing required header(s): "
            + ", ".join(missing)
        )
    return {field: actual[key] for field, key in expected.items()}


def read_ctc_info_workbook(xlsx_path: str | Path, images_dir: str | Path | None = None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Read the new workbook and return a descending-NUM article draft plus metadata.

    `sourceUrl` is the neutral canonical source-link field.  `tinyUrl` is also
    written with the same value solely for compatibility with existing page and
    Search code; historic articles retain their original TinyURL values.
    """
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
    worksheet = workbook.active
    columns = _column_map(worksheet)
    instruction = _text(worksheet.cell(2, 1).value)
    articles: list[dict[str, Any]] = []

    for row_number in range(DATA_START_ROW, worksheet.max_row + 1):
        values = {field: worksheet.cell(row_number, column).value for field, column in columns.items()}
        if all(_is_blank(value) for value in values.values()):
            continue

        # Some CTC Info workbooks retain a prose instruction row after the
        # article table.  It has no NUM, source URL, image name, or X-post URL,
        # so it is not an article row and must not be mistaken for one.
        if (
            _is_blank(values["num"])
            and _is_blank(values["sourceUrl"])
            and _is_blank(values["imageName"])
            and _is_blank(values["xPostUrl"])
        ):
            continue

        if _is_blank(values["num"]):
            raise ValueError(f"Row {row_number}: article data is present but Num (Column B) is blank")
        try:
            num = int(values["num"])
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Row {row_number}: Num must be an integer, got {values['num']!r}") from exc

        missing = [REQUIRED_HEADERS[field] for field, value in values.items() if field != "num" and _is_blank(value)]
        if missing:
            raise ValueError(f"Row {row_number}, NUM {num}: missing required field(s): {', '.join(missing)}")

        source_url = _text(values["sourceUrl"])
        image_name = _text(values["imageName"])
        article: dict[str, Any] = {
            "num": num,
            "headline": _text(values["headline"]),
            "sourceUrl": source_url,
            # Compatibility field for existing ArticleBlock/Search code.  It is
            # not a requirement that the URL be a TinyURL-domain address.
            "tinyUrl": source_url,
            "imageName": image_name,
            "xPostUrl": _text(values["xPostUrl"]),
            "linkOrigin": "source-url",
            "sourceRow": row_number,
        }
        if images_dir is not None:
            article["imagePath"] = str(Path(images_dir) / image_name)
        articles.append(article)

    workbook.close()
    if not articles:
        raise ValueError("No article rows found beginning at Row 5")

    nums = [article["num"] for article in articles]
    if len(nums) != len(set(nums)):
        raise ValueError("Duplicate NUM values found in CTC Info workbook")

    articles.sort(key=lambda article: article["num"], reverse=True)
    expected = list(range(articles[0]["num"], articles[-1]["num"] - 1, -1))
    if [article["num"] for article in articles] != expected:
        raise ValueError("NUM values must be contiguous; cannot calculate an unambiguous descending page order")

    metadata = {
        "headerRow": HEADER_ROW,
        "dataStartRow": DATA_START_ROW,
        "instruction": instruction,
        "articleCount": len(articles),
        "minimumNum": articles[-1]["num"],
        "maximumNum": articles[0]["num"],
    }
    return articles, metadata


def write_tagging_draft(xlsx_path: str | Path, images_dir: str | Path, output_json: str | Path) -> dict[str, Any]:
    """Create a post-ingestion draft for tag review; it never publishes a batch."""
    import json

    articles, metadata = read_ctc_info_workbook(xlsx_path, images_dir)
    for article in articles:
        article["tags"] = []
    Path(output_json).write_text(json.dumps(articles, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return metadata
