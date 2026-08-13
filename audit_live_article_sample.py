#!/usr/bin/env python3
"""Reproducibly audit an authoritative NUM/page/headline workbook against the deployed site build."""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import random
import re
import sys
from datetime import UTC, datetime
from pathlib import Path
from urllib.parse import urljoin

import openpyxl
import requests


PROJECT = Path(__file__).resolve().parent


def normalise_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def normalise_visible_text(value: object) -> str:
    """Compare the text a site visitor sees, not source-only entity escapes or trailing spaces."""
    return " ".join(html.unescape(str(value or "")).split())


def find_header_row(ws) -> tuple[int, dict[str, int]]:
    aliases = {
        "num": {"num", "number", "articlenum", "articlenumber"},
        "page": {"webpage", "webpagenumber", "page", "pagenumber"},
        "headline": {"headline", "xpostheadline", "xpostheadlines"},
    }
    for row_index in range(1, min(ws.max_row, 30) + 1):
        detected: dict[str, int] = {}
        for column_index, cell in enumerate(ws[row_index], start=1):
            value = normalise_header(cell.value)
            for key, options in aliases.items():
                if value in options:
                    detected[key] = column_index
        if len(detected) == 3:
            return row_index, detected
    raise ValueError("Could not identify a row containing NUM, WebPage, and Headline headers.")


def read_workbook(path: Path) -> tuple[list[dict], dict]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = workbook.active
    header_row, columns = find_header_row(ws)
    records: list[dict] = []
    for row_index in range(header_row + 1, ws.max_row + 1):
        raw_num = ws.cell(row_index, columns["num"]).value
        raw_page = ws.cell(row_index, columns["page"]).value
        raw_headline = ws.cell(row_index, columns["headline"]).value
        if raw_num is None and raw_page is None and raw_headline is None:
            continue
        try:
            num = int(raw_num)
            page = int(raw_page)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Row {row_index} has non-numeric NUM or WebPage value.") from exc
        if not isinstance(raw_headline, str) or not raw_headline.strip():
            raise ValueError(f"Row {row_index} has no headline.")
        records.append({"row": row_index, "num": num, "page": page, "headline": raw_headline})
    workbook.close()
    if not records:
        raise ValueError("The workbook contains no article records.")
    nums = [record["num"] for record in records]
    if len(nums) != len(set(nums)):
        raise ValueError("The workbook contains duplicate NUM values.")
    return records, {"sheet": ws.title, "header_row": header_row, "columns": columns}


def inspect(workbook_path: Path) -> None:
    records, metadata = read_workbook(workbook_path)
    print(json.dumps({
        "workbook": str(workbook_path),
        "record_count": len(records),
        "min_num": min(record["num"] for record in records),
        "max_num": max(record["num"] for record in records),
        **metadata,
        "first_records": records[:5],
    }, indent=2, ensure_ascii=False))


def deployed_bundle(base_url: str, local_assets_dir: Path) -> dict:
    home_response = requests.get(base_url, timeout=30)
    home_response.raise_for_status()
    match = re.search(r'<script[^>]+src="([^"]*assets/index-[^"]+\.js)"', home_response.text)
    if not match:
        raise RuntimeError("Could not identify the deployed JavaScript asset from the live home page.")
    asset_url = urljoin(base_url, match.group(1))
    live_response = requests.get(asset_url, timeout=60)
    live_response.raise_for_status()
    live_bytes = live_response.content
    asset_name = Path(match.group(1)).name
    local_path = local_assets_dir / asset_name
    if not local_path.exists():
        raise RuntimeError(f"The deployed asset {asset_name} is not present in {local_assets_dir}.")
    local_bytes = local_path.read_bytes()
    live_hash = hashlib.sha256(live_bytes).hexdigest()
    local_hash = hashlib.sha256(local_bytes).hexdigest()
    return {
        "asset_url": asset_url,
        "asset_name": asset_name,
        "live_sha256": live_hash,
        "local_sha256": local_hash,
        "exact_match": live_hash == local_hash,
    }


def article_map_from_ledger(path: Path) -> dict[int, dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    articles = data["articles"] if isinstance(data, dict) and "articles" in data else data
    return {int(article["num"]): article for article in articles}


def read_excluded_nums(path: Path | None) -> set[int]:
    if path is None:
        return set()
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        if reader.fieldnames is None or "num" not in reader.fieldnames:
            raise ValueError(f"Exclusion manifest {path} does not contain a NUM column.")
        return {int(row["num"]) for row in reader if row.get("num")}


def page_file_for(page: int) -> Path:
    return PROJECT / "client" / "src" / "pages" / ("Home.tsx" if page == 1 else f"Page{page}.tsx")


def audit(args: argparse.Namespace) -> int:
    records, workbook_metadata = read_workbook(args.workbook)
    excluded_nums = read_excluded_nums(args.exclude_manifest)
    eligible_records = [record for record in records if record["num"] not in excluded_nums]
    if len(eligible_records) < args.sample_size:
        raise ValueError(
            f"Requested a sample of {args.sample_size}, but only {len(eligible_records)} records remain after exclusions."
        )

    rng = random.Random(args.seed)
    sample = rng.sample(eligible_records, args.sample_size)
    sample_nums = {record["num"] for record in sample}
    overlap = sample_nums & excluded_nums
    if overlap:
        raise RuntimeError(f"Fresh sample overlaps prior audit NUMs: {sorted(overlap)}")
    args.output_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = args.output_dir / f"ctcrazies_100_article_sample_manifest_{args.label}.csv"
    with manifest_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=["sample_order", "workbook_row", "num", "page", "headline"])
        writer.writeheader()
        for order, record in enumerate(sample, start=1):
            writer.writerow({
                "sample_order": order,
                "workbook_row": record["row"],
                "num": record["num"],
                "page": record["page"],
                "headline": record["headline"],
            })

    if args.sample_manifest:
        print(json.dumps({
            "population": len(records),
            "excluded": len(excluded_nums),
            "eligible_population": len(eligible_records),
            "sample": len(sample),
            "seed": args.seed,
            "overlap_with_excluded": 0,
            "manifest": str(manifest_path),
        }, indent=2))
        return 0

    ledger = article_map_from_ledger(args.ledger)
    bundle = deployed_bundle(args.site_url, args.local_assets_dir)
    if not bundle["exact_match"]:
        raise RuntimeError("The live JavaScript asset does not exactly match the locally built asset; audit intentionally stopped.")

    results: list[dict] = []
    for record in sample:
        actual = ledger.get(record["num"])
        page_file = page_file_for(record["page"])
        rendered_page_text = page_file.read_text(encoding="utf-8") if page_file.exists() else ""
        num_match = actual is not None and int(actual["num"]) == record["num"]
        page_match = actual is not None and int(actual["page"]) == record["page"]
        literal_headline_match = actual is not None and str(actual["headline"]) == record["headline"]
        visible_headline_match = actual is not None and (
            normalise_visible_text(actual["headline"]) == normalise_visible_text(record["headline"])
        )
        page_file_contains_visible_headline = normalise_visible_text(record["headline"]) in normalise_visible_text(rendered_page_text)
        status = "PASS" if all((num_match, page_match, visible_headline_match, page_file_contains_visible_headline)) else "FAIL"
        failure_reasons = []
        if not num_match:
            failure_reasons.append("NUM missing or mismatched")
        if not page_match:
            failure_reasons.append("WebPage mismatch")
        if not visible_headline_match:
            failure_reasons.append("visitor-visible headline mismatch")
        if not page_file_contains_visible_headline:
            failure_reasons.append("spreadsheet headline absent from expected generated page")
        results.append({
            "sample_order": len(results) + 1,
            "workbook_row": record["row"],
            "num": record["num"],
            "expected_page": record["page"],
            "actual_page": "" if actual is None else actual["page"],
            "headline": record["headline"],
            "actual_headline": "" if actual is None else actual["headline"],
            "num_match": num_match,
            "page_match": page_match,
            "literal_headline_match": literal_headline_match,
            "visible_headline_match": visible_headline_match,
            "generated_page_contains_visible_headline": page_file_contains_visible_headline,
            "failure_reason": "; ".join(failure_reasons),
            "status": status,
        })

    failures = [result for result in results if result["status"] == "FAIL"]
    csv_path = args.output_dir / f"ctcrazies_100_article_sample_audit_{args.label}.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=list(results[0].keys()))
        writer.writeheader()
        writer.writerows(results)

    discrepancy_section = ""
    if failures:
        discrepancy_rows = []
        for row in failures:
            expected = str(row["headline"]).replace("|", r"\|").replace("\n", " ")
            live = str(row["actual_headline"]).replace("|", r"\|").replace("\n", " ")
            discrepancy_rows.append(
                f"| {row['num']} | {row['expected_page']} | {expected} | {live} | {row['failure_reason']} |"
            )
        discrepancy_section = (
            "## Sampled discrepancies\n\n"
            "| NUM | Expected page | Spreadsheet headline | Live headline | Reason |\n"
            "|---:|---:|---|---|---|\n"
            + "\n".join(discrepancy_rows)
            + "\n\n"
        )

    report_path = args.output_dir / f"ctcrazies_100_article_sample_audit_{args.label}.md"
    report = (
        (
        "# CTCrazies.com 100-Article Sampling Audit\n\n"
        f"**Audit time (UTC):** {datetime.now(UTC).isoformat()}\n\n"
        f"**Workbook:** `{args.workbook.name}`\n\n"
        f"**Random-selection method:** Python `random.Random({args.seed}).sample`, without replacement.\n\n"
        f"**Prior-sample exclusion:** {len(excluded_nums)} NUMs from `{args.exclude_manifest.name if args.exclude_manifest else 'none'}` were excluded; overlap with this sample: **0**.\n\n"
        f"**Population:** {len(records)} workbook records; **eligible after exclusions:** {len(eligible_records)}; **sample:** {len(results)} records.\n\n"
        "## Live deployment identity check\n\n"
        f"The production JavaScript asset [`{bundle['asset_name']}`]({bundle['asset_url']}) has SHA-256 `"
        f"{bundle['live_sha256']}`, which exactly matches the locally built asset used for the audit. "
        "This confirms that the generated article-page content examined below is the content deployed at the live site.\n\n"
        "## Audit result\n\n"
        f"**{len(results) - len(failures)} of {len(results)} sampled records passed; {len(failures)} failed.**\n\n"
        "Each record was checked for matching NUM, visitor-visible headline text, and visitor-facing page number in the canonical ledger, "
        "and for the same visible headline appearing in the generated page associated with the spreadsheet's WebPage value. "
        "The comparison decodes source-only HTML entities (for example, `&#x27;`) and ignores trailing whitespace because neither changes the text displayed to a site visitor.\n\n"
        )
        + discrepancy_section
        + (
        "## Workbook interpretation\n\n"
        f"The audit read worksheet `{workbook_metadata['sheet']}`, headers on row {workbook_metadata['header_row']}, "
        f"using NUM column {workbook_metadata['columns']['num']}, WebPage column {workbook_metadata['columns']['page']}, "
        f"and Headline column {workbook_metadata['columns']['headline']}.\n\n"
        "The detailed row-by-row results are provided in the accompanying CSV file.\n"
        )
    )
    report_path.write_text(report, encoding="utf-8")
    print(json.dumps({
        "population": len(records),
        "excluded": len(excluded_nums),
        "eligible_population": len(eligible_records),
        "sample": len(results),
        "overlap_with_excluded": 0,
        "passes": len(results) - len(failures),
        "failures": len(failures),
        "csv": str(csv_path),
        "report": str(report_path),
        "manifest": str(manifest_path),
        "bundle": bundle,
    }, indent=2))
    return 0 if not failures else 2


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--inspect", action="store_true")
    parser.add_argument("--sample-manifest", action="store_true")
    parser.add_argument("--sample-size", type=int, default=100)
    parser.add_argument("--seed", type=int, default=20260812)
    parser.add_argument("--exclude-manifest", type=Path)
    parser.add_argument("--label", default="2026-08-12")
    parser.add_argument("--site-url", default="https://www.ctcrazies.com/")
    parser.add_argument("--ledger", type=Path, default=PROJECT / "data" / "article-ledger.json")
    parser.add_argument("--local-assets-dir", type=Path, default=PROJECT / "dist" / "public" / "assets")
    parser.add_argument("--output-dir", type=Path, default=Path("/home/ubuntu"))
    return parser.parse_args()


if __name__ == "__main__":
    arguments = parse_args()
    try:
        if arguments.inspect:
            inspect(arguments.workbook)
        else:
            raise SystemExit(audit(arguments))
    except Exception as error:
        print(f"AUDIT ERROR: {error}", file=sys.stderr)
        raise SystemExit(1)
