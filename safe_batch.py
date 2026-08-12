"""CTCrazies safe batch system.

This is the only supported publication-preparation tool for future batches.  It
uses a repository-tracked article ledger as the single source of truth, derives
visitor-facing page placement from descending NUM order, and creates both
Search.tsx and the two typed tag indexes from that same ordered ledger.

It deliberately does not deploy to Cloudflare.  Deployment is allowed only
after `verify_safe_site.py` exits successfully.
"""

from __future__ import annotations

import argparse
import copy
import html
import json
import re
import shutil
import sys
import tempfile
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from ctc_info_xlsx import read_ctc_info_workbook, write_tagging_draft


PROJECT = Path(__file__).resolve().parent
PAGES = PROJECT / "client" / "src" / "pages"
SEARCH = PAGES / "Search.tsx"
APP = PROJECT / "client" / "src" / "App.tsx"
TAG_SRC = PROJECT / "client" / "src" / "data" / "tag-index.json"
TAG_PUBLIC = PROJECT / "client" / "public" / "tag-index.json"
LEDGER = PROJECT / "data" / "article-ledger.json"
ARTICLE_IMAGES_CDN = "https://cdn.jsdelivr.net/gh/chrisgarson/CTCraziesBlog@main/article-images"
ARTICLES_PER_PAGE = 20


def normalize(value: str) -> str:
    return " ".join(html.unescape(value or "").replace("\xa0", " ").split()).casefold()


def article_key(article: dict[str, Any]) -> tuple[str, str]:
    """A stable identity that preserves legacy TinyURL links and new Source URLs."""
    return (str(article.get("sourceUrl") or article.get("tinyUrl") or ""), str(article.get("xPostUrl") or ""))


def page_path(page: int, root: Path = PAGES) -> Path:
    return root / ("Home.tsx" if page == 1 else f"Page{page}.tsx")


def extract_blocks(content: str) -> list[str]:
    return re.findall(r"<ArticleBlock\b.*?/>", content, re.DOTALL)


def jsx_attr(block: str, attribute: str) -> str:
    match = re.search(rf'\b{re.escape(attribute)}="((?:[^"\\]|\\.)*)"', block, re.DOTALL)
    return html.unescape(match.group(1)) if match else ""


def jsx_tags(block: str) -> list[str]:
    match = re.search(r"tags=\{\[([^\]]*)\]\}", block, re.DOTALL)
    if not match:
        return []
    return [html.unescape(value) for value in re.findall(r'"((?:[^"\\]|\\.)*)"', match.group(1))]


def current_page_articles(project: Path = PROJECT) -> list[dict[str, Any]]:
    """Extract physical article blocks in visitor-facing page order."""
    pages_dir = project / "client" / "src" / "pages"
    numbers = [1] + sorted(
        int(match.group(1))
        for path in pages_dir.glob("Page*.tsx")
        if (match := re.fullmatch(r"Page(\d+)\.tsx", path.name))
    )
    result: list[dict[str, Any]] = []
    for page in numbers:
        path = page_path(page, pages_dir)
        blocks = extract_blocks(path.read_text(encoding="utf-8"))
        if len(blocks) != ARTICLES_PER_PAGE:
            raise ValueError(f"{path.name} contains {len(blocks)} articles; expected {ARTICLES_PER_PAGE}")
        for position, block in enumerate(blocks, start=1):
            result.append({
                "page": page,
                "position": position,
                "headline": jsx_attr(block, "headline"),
                "sourceUrl": jsx_attr(block, "tinyUrl"),
                "xPostUrl": jsx_attr(block, "xPostUrl"),
                "imageUrl": jsx_attr(block, "imageSrc"),
                "tags": jsx_tags(block),
            })
    return result


def find_articles_array(content: str) -> tuple[int, int]:
    start = content.find("const articles = [")
    if start < 0:
        raise ValueError("Search.tsx does not contain the articles array")
    bracket = content.find("[", start)
    depth, in_string, escaped = 0, False, False
    for index in range(bracket, len(content)):
        char = content[index]
        if in_string:
            if escaped:
                escaped = False
            elif char == "\\":
                escaped = True
            elif char == '"':
                in_string = False
            continue
        if char == '"':
            in_string = True
        elif char == "[":
            depth += 1
        elif char == "]":
            depth -= 1
            if depth == 0:
                return start, index + 1
    raise ValueError("Search.tsx articles array is unterminated")


def search_records(project: Path = PROJECT) -> dict[tuple[str, str], dict[str, Any]]:
    """Read only fields needed to carry forward the historic batch date."""
    content = (project / "client" / "src" / "pages" / "Search.tsx").read_text(encoding="utf-8")
    start, end = find_articles_array(content)
    section = content[start:end]
    records: dict[tuple[str, str], dict[str, Any]] = {}
    array_start = section.find("[")
    try:
        # Safe-generated records are valid JSON and preserve numeric NUM/page
        # values exactly. Prefer this parser to indentation-sensitive regexes.
        entries = json.loads(section[array_start:])
        for entry in entries:
            source, xpost = str(entry.get("tinyUrl", "")), str(entry.get("xPostUrl", ""))
            if source or xpost:
                records[(source, xpost)] = {
                    "headline": str(entry.get("headline", "")),
                    "batchDate": str(entry.get("batchDate", "")),
                    "page": entry.get("page"),
                }
        return records
    except json.JSONDecodeError:
        pass
    # Legacy records used non-JSON TypeScript syntax; retain a conservative
    # fallback solely for migration/bootstrap support.
    for block in re.findall(r"\{.*?\n  \}", section, re.DOTALL):
        def field(name: str) -> str:
            match = re.search(rf'"{re.escape(name)}"\s*:\s*"((?:[^"\\]|\\.)*)"', block)
            return json.loads('"' + match.group(1) + '"') if match else ""
        source, xpost = field("tinyUrl"), field("xPostUrl")
        if source or xpost:
            page_match = re.search(r'"page"\s*:\s*(\d+)', block)
            records[(source, xpost)] = {
                "headline": field("headline"),
                "batchDate": field("batchDate"),
                "page": int(page_match.group(1)) if page_match else None,
            }
    return records


def _normal_header(value: Any) -> str:
    return "".join(character.lower() for character in str(value or "") if character.isalnum())


def authoritative_records(xlsx_path: str | Path) -> list[dict[str, Any]]:
    """Read an existing all-article NUM/WebPage mapping from a flexible workbook."""
    import openpyxl

    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    headers: dict[str, int] | None = None
    header_row = 0
    for row_number in range(1, min(sheet.max_row, 30) + 1):
        candidate = {_normal_header(sheet.cell(row_number, col).value): col for col in range(1, sheet.max_column + 1)}
        if "num" in candidate and any(key in candidate for key in ("webpage", "webpagenumber")):
            headers, header_row = candidate, row_number
            break
    if not headers:
        raise ValueError("Authoritative workbook must contain NUM and WebPage# headers")
    page_col = headers.get("webpage") or headers.get("webpagenumber")
    headline_col = headers.get("xpostheadline") or headers.get("headline")
    if not headline_col:
        raise ValueError("Authoritative workbook must contain an X-Post Headline column")
    records = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        num, page = sheet.cell(row_number, headers["num"]).value, sheet.cell(row_number, page_col).value
        headline = sheet.cell(row_number, headline_col).value
        if num is None and page is None and headline is None:
            continue
        try:
            records.append({"num": int(num), "page": int(page), "headline": str(headline or "")})
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Authoritative workbook row {row_number} has invalid NUM/WebPage#") from exc
    workbook.close()
    return records


def tag_metadata(project: Path = PROJECT) -> dict[str, dict[str, Any]]:
    data = json.loads((project / "client" / "src" / "data" / "tag-index.json").read_text(encoding="utf-8"))
    metadata: dict[str, dict[str, Any]] = {}
    for tag, entry in data.items():
        if not isinstance(entry, dict) or "type" not in entry:
            raise ValueError("Current tag index is not the required typed schema")
        metadata[tag] = {
            "type": entry["type"],
            "keywords": list(entry.get("keywords", [])),
        }
    return metadata


def existing_tag_associations(project: Path = PROJECT) -> dict[tuple[str, str], set[str]]:
    """Collect tag associations already approved in the typed tag database.

    During the one-time ledger bootstrap, the tag database may contain an
    approved association that is absent from a legacy page component.  The
    canonical ledger must retain that association rather than silently delete
    it during its first typed-index rebuild.
    """
    data = json.loads((project / "client" / "src" / "data" / "tag-index.json").read_text(encoding="utf-8"))
    associations: dict[tuple[str, str], set[str]] = defaultdict(set)
    for tag, entry in data.items():
        for article in entry.get("articles", []):
            associations[article_key(article)].add(tag)
    return associations


def assert_descending_sequence(articles: list[dict[str, Any]]) -> None:
    nums = [int(article["num"]) for article in articles]
    if not nums:
        raise ValueError("Article ledger is empty")
    expected = list(range(nums[0], nums[-1] - 1, -1))
    if nums != expected:
        raise ValueError("Article NUMs are not a unique, contiguous descending sequence")
    if len(articles) % ARTICLES_PER_PAGE:
        raise ValueError(f"Article total must be divisible by {ARTICLES_PER_PAGE}")


def assign_pages(articles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ordered = sorted(copy.deepcopy(articles), key=lambda article: int(article["num"]), reverse=True)
    assert_descending_sequence(ordered)
    for index, article in enumerate(ordered):
        article["page"] = index // ARTICLES_PER_PAGE + 1
    return ordered


def bootstrap_ledger(authoritative_xlsx: str | Path, output: Path = LEDGER) -> dict[str, Any]:
    """Create the one-time canonical ledger from the currently verified website."""
    physical = current_page_articles()
    authoritative = sorted(authoritative_records(authoritative_xlsx), key=lambda item: (item["page"], -item["num"]))
    if len(physical) != len(authoritative):
        raise ValueError(f"Physical site has {len(physical)} articles; authoritative workbook has {len(authoritative)}")
    for source, expected in zip(physical, authoritative):
        if source["page"] != expected["page"]:
            raise ValueError(f"Current Page {source['page']} does not match authoritative Page {expected['page']}")
    dates = search_records()
    approved_associations = existing_tag_associations()
    articles: list[dict[str, Any]] = []
    for source, expected in zip(physical, authoritative):
        key = article_key(source)
        retained_tags = sorted(set(source["tags"]) | approved_associations.get(key, set()), key=str.casefold)
        search_entry = dates.get(key, {})
        articles.append({
            "num": expected["num"],
            "headline": search_entry.get("headline") or source["headline"],
            "sourceUrl": source["sourceUrl"],
            "xPostUrl": source["xPostUrl"],
            "imageUrl": source["imageUrl"],
            "tags": retained_tags,
            "batchDate": search_entry.get("batchDate", ""),
            "linkOrigin": "legacy-tinyurl" if "tinyurl." in source["sourceUrl"].lower() else "source-url",
        })
    articles = assign_pages(articles)
    ledger = {
        "schemaVersion": 1,
        "articlesPerPage": ARTICLES_PER_PAGE,
        "tagMetadata": tag_metadata(),
        "articles": articles,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(ledger, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return ledger


def load_ledger(path: Path = LEDGER) -> dict[str, Any]:
    if not path.exists():
        raise ValueError(f"Canonical article ledger does not exist: {path}")
    ledger = json.loads(path.read_text(encoding="utf-8"))
    if ledger.get("schemaVersion") != 1 or ledger.get("articlesPerPage") != ARTICLES_PER_PAGE:
        raise ValueError("Unsupported article-ledger schema")
    ledger["articles"] = assign_pages(ledger.get("articles", []))
    return ledger


def write_ledger(ledger: dict[str, Any], path: Path = LEDGER) -> None:
    path.write_text(json.dumps(ledger, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def validate_batch(ledger: dict[str, Any], batch: list[dict[str, Any]], xlsx_path: str | Path | None = None) -> list[dict[str, Any]]:
    if not batch or len(batch) % ARTICLES_PER_PAGE:
        raise ValueError("Incoming batch must contain a positive multiple of 20 articles")
    existing = ledger["articles"]
    maximum = max(article["num"] for article in existing)
    normalized: list[dict[str, Any]] = []
    for article in batch:
        item = copy.deepcopy(article)
        item["sourceUrl"] = item.get("sourceUrl") or item.get("tinyUrl") or ""
        item["tinyUrl"] = item["sourceUrl"]
        required = ["num", "headline", "sourceUrl", "imageName", "xPostUrl", "tags"]
        missing = [field for field in required if not str(item.get(field, "")).strip() and field != "tags"]
        if missing or not isinstance(item.get("tags"), list) or not item["tags"]:
            raise ValueError(f"NUM {item.get('num', '?')}: missing required published-batch data")
        image_path = item.get("imagePath")
        if image_path and not Path(image_path).exists():
            raise ValueError(f"NUM {item['num']}: image file not found: {image_path}")
        if not item.get("imageUrl"):
            item["imageUrl"] = f"{ARTICLE_IMAGES_CDN}/{item['imageName']}"
        normalized.append(item)
    normalized.sort(key=lambda item: int(item["num"]), reverse=True)
    nums = [int(item["num"]) for item in normalized]
    expected = list(range(maximum + len(normalized), maximum, -1))
    if nums != expected:
        raise ValueError(f"Incoming NUMs must be exactly {expected[0]} through {expected[-1]} in descending order")
    unknown_tags = sorted({tag for item in normalized for tag in item["tags"] if tag not in ledger["tagMetadata"]})
    if unknown_tags:
        raise ValueError("New tags require explicit metadata approval before publication: " + ", ".join(unknown_tags))
    if xlsx_path:
        xlsx_articles, _ = read_ctc_info_workbook(xlsx_path)
        xlsx_by_num = {item["num"]: item for item in xlsx_articles}
        for item in normalized:
            xlsx_item = xlsx_by_num.get(item["num"])
            if not xlsx_item or item["headline"] != xlsx_item["headline"] or item["sourceUrl"] != xlsx_item["sourceUrl"]:
                raise ValueError(f"NUM {item['num']}: batch does not exactly match the CTC Info workbook")
    return normalized


def combine_batch(ledger: dict[str, Any], batch: list[dict[str, Any]], batch_date: str) -> dict[str, Any]:
    combined = copy.deepcopy(ledger)
    for item in batch:
        item["batchDate"] = batch_date
        item["linkOrigin"] = "source-url"
        item.pop("imageName", None)
        item.pop("imagePath", None)
        item.pop("tinyUrl", None)
        item.pop("sourceRow", None)
        combined["articles"].append(item)
    combined["articles"] = assign_pages(combined["articles"])
    return combined


def jsx_escape(value: str) -> str:
    return html.escape(value, quote=True)


def render_article(article: dict[str, Any]) -> str:
    tags = json.dumps(article["tags"], ensure_ascii=False)
    return "\n".join([
        "      <ArticleBlock",
        f'        headline="{jsx_escape(article["headline"])}"',
        f'        tinyUrl="{jsx_escape(article["sourceUrl"])}"',
        f'        xPostUrl="{jsx_escape(article["xPostUrl"])}"',
        f'        imageSrc="{jsx_escape(article["imageUrl"])}"',
        f"        tags={{{tags}}}",
        "      />",
    ])


def render_page(page: int, articles: list[dict[str, Any]], total_pages: int, batch_date: str) -> str:
    name = "Home" if page == 1 else f"Page{page}"
    blocks = "\n".join(render_article(article) for article in articles)
    stats = ""
    if page == 1:
        stats = "\n".join([
            "      {/* SITE STATS — generated by safe_batch.py. */}",
            "      <p className=\"text-sm text-gray-500 mt-2\" style={{fontFamily: 'Roboto Slab, serif'}}>",
            f"        Last updated: <strong style={{{{ color: '#555' }}}}>{jsx_escape(batch_date)}</strong> &nbsp;|&nbsp; Total articles: <strong style={{{{ color: '#555' }}}}>{total_pages * ARTICLES_PER_PAGE}</strong>",
            "      </p>",
        ])
    return "\n".join([
        "import ArticleBlock from '../components/ArticleBlock';",
        "import PageHeader from '../components/PageHeader';",
        "import Pagination from '../components/Pagination';",
        "",
        f"export default function {name}() {{",
        "  return (",
        '    <div className="max-w-4xl mx-auto px-4 py-8">',
        "      <PageHeader />",
        '      <div className="space-y-12">',
        blocks,
        "      </div>",
        f"      <Pagination currentPage={{{page}}} totalPages={{{total_pages}}} />",
        stats,
        "    </div>",
        "  );",
        "}",
        "",
    ])


def render_app(total_pages: int) -> str:
    imports = ['import { Switch, Route } from "wouter";', 'import Home from "./pages/Home";']
    imports += [f'import Page{page} from "./pages/Page{page}";' for page in range(2, total_pages + 1)]
    imports += ['import Search from "./pages/Search";', 'import TagResults from "./pages/TagResults";', 'import TagsIndex from "./pages/TagsIndex";']
    routes = ['        <Route path="/" component={Home} />']
    routes += [f'        <Route path="/page{page}" component={{Page{page}}} />' for page in range(2, total_pages + 1)]
    routes += ['        <Route path="/search" component={Search} />', '        <Route path="/tags" component={TagsIndex} />', '        <Route path="/tag/:tag" component={TagResults} />']
    return "\n".join(imports + ["", "function App() {", "  return (", "    <>", "      <Switch>", *routes, "      </Switch>", "    </>", "  );", "}", "", "export default App;", ""])


def render_search(articles: list[dict[str, Any],], original: str) -> str:
    entries = []
    for article in articles:
        entries.append(json.dumps({
            "num": article["num"],
            "headline": article["headline"],
            "tinyUrl": article["sourceUrl"],
            "xPostUrl": article["xPostUrl"],
            "imageUrl": article["imageUrl"],
            "tags": article["tags"],
            "page": article["page"],
            "batchDate": article.get("batchDate", ""),
        }, ensure_ascii=False, indent=4))
    array = "const articles = [\n" + ",\n".join(entries) + "\n]"
    start, end = find_articles_array(original)
    return original[:start] + array + original[end:] 


def build_tag_index(ledger: dict[str, Any]) -> dict[str, Any]:
    metadata = copy.deepcopy(ledger["tagMetadata"])
    index = {tag: {"type": item["type"], "articles": [], **({"keywords": item["keywords"]} if item.get("keywords") else {})} for tag, item in metadata.items()}
    for article in ledger["articles"]:
        for tag in article["tags"]:
            if tag not in index:
                raise ValueError(f"NUM {article['num']} references unapproved tag {tag!r}")
            index[tag]["articles"].append({
                "headline": article["headline"],
                "tinyUrl": article["sourceUrl"],
                "xPostUrl": article["xPostUrl"],
                "imageUrl": article["imageUrl"],
                "page": article["page"],
                "batchDate": article.get("batchDate", ""),
            })
    for entry in index.values():
        entry["articles"].sort(key=lambda article: (article["page"], article["headline"].casefold()))
    return dict(sorted(index.items(), key=lambda item: item[0].casefold()))


def write_project_from_ledger(ledger: dict[str, Any], batch_date: str, project: Path = PROJECT) -> None:
    """Generate every derived file only after the ledger has passed all checks."""
    total_pages = len(ledger["articles"]) // ARTICLES_PER_PAGE
    pages_dir = project / "client" / "src" / "pages"
    staged = Path(tempfile.mkdtemp(prefix="ctc_safe_batch_"))
    try:
        groups: dict[int, list[dict[str, Any]]] = defaultdict(list)
        for article in ledger["articles"]:
            groups[article["page"]].append(article)
        for page in range(1, total_pages + 1):
            if len(groups[page]) != ARTICLES_PER_PAGE:
                raise ValueError(f"Calculated Page {page} has {len(groups[page])} articles")
            (staged / ("Home.tsx" if page == 1 else f"Page{page}.tsx")).write_text(render_page(page, groups[page], total_pages, batch_date), encoding="utf-8")
        search_text = render_search(ledger["articles"], (pages_dir / "Search.tsx").read_text(encoding="utf-8"))
        (staged / "Search.tsx").write_text(search_text, encoding="utf-8")
        (staged / "App.tsx").write_text(render_app(total_pages), encoding="utf-8")
        tag_text = json.dumps(build_tag_index(ledger), ensure_ascii=False, indent=2) + "\n"
        (staged / "tag-index.json").write_text(tag_text, encoding="utf-8")

        # Replace page files only once all staged files are complete.
        for existing in pages_dir.glob("Page*.tsx"):
            existing.unlink()
        for source in staged.glob("Page*.tsx"):
            shutil.copy2(source, pages_dir / source.name)
        shutil.copy2(staged / "Home.tsx", pages_dir / "Home.tsx")
        shutil.copy2(staged / "Search.tsx", pages_dir / "Search.tsx")
        shutil.copy2(staged / "App.tsx", project / "client" / "src" / "App.tsx")
        shutil.copy2(staged / "tag-index.json", project / "client" / "src" / "data" / "tag-index.json")
        shutil.copy2(staged / "tag-index.json", project / "client" / "public" / "tag-index.json")
    finally:
        shutil.rmtree(staged, ignore_errors=True)


def prepare_command(args: argparse.Namespace) -> None:
    metadata = write_tagging_draft(args.xlsx, args.images_dir, args.output)
    print(json.dumps(metadata, ensure_ascii=False, indent=2))


def bootstrap_command(args: argparse.Namespace) -> None:
    ledger = bootstrap_ledger(args.authoritative_xlsx, Path(args.output))
    print(f"Bootstrapped canonical ledger: {len(ledger['articles'])} articles, {len(ledger['tagMetadata'])} tags")


def validate_command(args: argparse.Namespace) -> None:
    ledger = load_ledger(Path(args.ledger))
    batch = json.loads(Path(args.batch_json).read_text(encoding="utf-8"))
    batch = validate_batch(ledger, batch, args.xlsx)
    combined = combine_batch(ledger, batch, args.batch_date)
    print(f"VALID: {len(batch)} incoming articles; {len(combined['articles'])} total; {len(combined['articles']) // ARTICLES_PER_PAGE} pages")


def apply_command(args: argparse.Namespace) -> None:
    ledger_path = Path(args.ledger)
    ledger = load_ledger(ledger_path)
    batch = json.loads(Path(args.batch_json).read_text(encoding="utf-8"))
    batch = validate_batch(ledger, batch, args.xlsx)
    combined = combine_batch(ledger, batch, args.batch_date)
    write_project_from_ledger(combined, args.batch_date)
    write_ledger(combined, ledger_path)
    print(f"APPLIED: {len(batch)} articles; {len(combined['articles'])} total; {len(combined['articles']) // ARTICLES_PER_PAGE} pages")
    print("Run verify_safe_site.py and pnpm run build before GitHub commit or Cloudflare deployment.")


def sync_indexes_command(args: argparse.Namespace) -> None:
    """Synchronize both typed tag indexes from the canonical ledger without rewriting pages."""
    ledger = load_ledger(Path(args.ledger))
    tag_text = json.dumps(build_tag_index(ledger), ensure_ascii=False, indent=2) + "\n"
    TAG_SRC.write_text(tag_text, encoding="utf-8")
    TAG_PUBLIC.write_text(tag_text, encoding="utf-8")
    print(f"SYNCHRONIZED: both typed tag indexes now reflect {len(ledger['articles'])} canonical articles")


def declare_tag(ledger: dict[str, Any], tag: str, tag_type: str, keywords: list[str] | None = None) -> dict[str, Any]:
    """Add user-approved metadata before a new tag can be used in a batch."""
    tag = tag.strip()
    if not tag:
        raise ValueError("Tag name cannot be blank")
    existing = ledger["tagMetadata"].get(tag)
    retained_keywords = list(existing.get("keywords", [])) if existing else []
    for keyword in keywords or []:
        if keyword not in retained_keywords:
            retained_keywords.append(keyword)
    if existing and existing["type"] != tag_type:
        raise ValueError(f"Tag {tag!r} already exists as type {existing['type']!r}")
    ledger["tagMetadata"][tag] = {"type": tag_type, "keywords": retained_keywords}
    return ledger


def declare_tag_command(args: argparse.Namespace) -> None:
    """Record a user-approved new person or topic tag before it can be used in a batch."""
    ledger_path = Path(args.ledger)
    ledger = declare_tag(load_ledger(ledger_path), args.tag, args.type, args.keyword)
    write_ledger(ledger, ledger_path)
    sync_indexes_command(argparse.Namespace(ledger=str(ledger_path)))
    print(f"DECLARED: {args.tag.strip()} ({args.type})")


def rename_tag(ledger: dict[str, Any], old_tag: str, new_tag: str) -> dict[str, Any]:
    """Rename one approved tag while retaining its type, keywords, and article associations."""
    old_tag, new_tag = old_tag.strip(), new_tag.strip()
    if not old_tag or not new_tag:
        raise ValueError("Old and new tag names cannot be blank")
    if old_tag not in ledger["tagMetadata"]:
        raise ValueError(f"Active tag {old_tag!r} does not exist")
    if new_tag in ledger["tagMetadata"]:
        raise ValueError(f"Destination tag {new_tag!r} already exists; use a consolidation instead")
    ledger["tagMetadata"][new_tag] = ledger["tagMetadata"].pop(old_tag)
    for article in ledger["articles"]:
        if old_tag in article["tags"]:
            article["tags"] = [new_tag if tag == old_tag else tag for tag in article["tags"]]
            article["tags"] = list(dict.fromkeys(article["tags"]))
    return ledger


def rename_tag_command(args: argparse.Namespace) -> None:
    """Rename an approved tag and regenerate every derived source file safely."""
    ledger_path = Path(args.ledger)
    ledger = rename_tag(load_ledger(ledger_path), args.old_tag, args.new_tag)
    latest_date = ledger["articles"][0].get("batchDate") or date.today().isoformat()
    write_project_from_ledger(ledger, latest_date)
    write_ledger(ledger, ledger_path)
    print(f"RENAMED: {args.old_tag!r} → {args.new_tag!r}; run verify_safe_site.py before publishing")


def render_current_command(args: argparse.Namespace) -> None:
    """Regenerate derived source files from the existing canonical ledger without metadata changes."""
    ledger = load_ledger(Path(args.ledger))
    latest_date = ledger["articles"][0].get("batchDate") or date.today().isoformat()
    write_project_from_ledger(ledger, latest_date)
    print("RENDERED: all pages, Search.tsx, App.tsx, and both typed tag indexes regenerated from the ledger")


def correct_ledger_headline_command(args: argparse.Namespace) -> None:
    """Apply a verified headline correction to one canonical article by its immutable X-post URL."""
    ledger_path = Path(args.ledger)
    ledger = load_ledger(ledger_path)
    matches = [article for article in ledger["articles"] if article["xPostUrl"] == args.xpost_url]
    if len(matches) != 1:
        raise ValueError("Expected exactly one canonical article for the supplied X-post URL")
    matches[0]["headline"] = args.headline
    write_project_from_ledger(ledger, ledger["articles"][0].get("batchDate") or date.today().isoformat())
    write_ledger(ledger, ledger_path)
    print(f"CORRECTED HEADLINE: NUM {matches[0]['num']}; run verify_safe_site.py before publishing")


def main() -> None:
    parser = argparse.ArgumentParser(description="Safe CTCrazies batch system")
    commands = parser.add_subparsers(dest="command", required=True)
    prepare = commands.add_parser("prepare", help="Read a CTC Info workbook into a tag-review draft")
    prepare.add_argument("xlsx")
    prepare.add_argument("images_dir")
    prepare.add_argument("output")
    prepare.set_defaults(func=prepare_command)
    bootstrap = commands.add_parser("bootstrap", help="Create ledger from current verified pages and authoritative mapping")
    bootstrap.add_argument("authoritative_xlsx")
    bootstrap.add_argument("--output", default=str(LEDGER))
    bootstrap.set_defaults(func=bootstrap_command)
    for name, handler in (("validate", validate_command), ("apply", apply_command)):
        command = commands.add_parser(name)
        command.add_argument("batch_date")
        command.add_argument("batch_json")
        command.add_argument("--xlsx")
        command.add_argument("--ledger", default=str(LEDGER))
        command.set_defaults(func=handler)
    sync = commands.add_parser("sync-indexes", help="Synchronize both typed tag indexes from the canonical ledger")
    sync.add_argument("--ledger", default=str(LEDGER))
    sync.set_defaults(func=sync_indexes_command)
    declare = commands.add_parser("declare-tag", help="Add a user-approved person or topic tag to ledger metadata")
    declare.add_argument("tag")
    declare.add_argument("--type", required=True, choices=("person", "topic"))
    declare.add_argument("--keyword", action="append", help="Optional retired/alias keyword; may be specified more than once")
    declare.add_argument("--ledger", default=str(LEDGER))
    declare.set_defaults(func=declare_tag_command)
    rename = commands.add_parser("rename-tag", help="Rename one approved tag across the canonical ledger and derived files")
    rename.add_argument("old_tag")
    rename.add_argument("new_tag")
    rename.add_argument("--ledger", default=str(LEDGER))
    rename.set_defaults(func=rename_tag_command)
    render_current = commands.add_parser("render-current", help="Regenerate all derived source files from the canonical ledger")
    render_current.add_argument("--ledger", default=str(LEDGER))
    render_current.set_defaults(func=render_current_command)
    headline = commands.add_parser("correct-headline", help="Apply a verified title correction to one canonical article")
    headline.add_argument("xpost_url")
    headline.add_argument("headline")
    headline.add_argument("--ledger", default=str(LEDGER))
    headline.set_defaults(func=correct_ledger_headline_command)
    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(f"SAFE BATCH FAILED: {error}", file=sys.stderr)
        sys.exit(1)
